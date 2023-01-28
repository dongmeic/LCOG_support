import pandas as pd
import geopandas as gpd
import os
import requests
from bs4 import BeautifulSoup
from homeless import *
import openpyxl
from win32com.client import Dispatch
import numpy as np

outpath = r'G:\projects\UtilityDistricts\eweb\DrinkingWater\IllegalCampCoordination\Recieved'
path = outpath + '\\IllegalCampNotification_pro'

dat = pd.read_excel(path+'\\most_recent.xlsx')
cols2drop = ['OBJECTID', 'Join_Count', 'Unruly_inhabitants']
dat = dat.drop(cols2drop, axis=1)
dat.columns = list(map(lambda x: x.capitalize(), dat.columns))
dat.rename(columns={'Ownname': 'Owner_name', 'Addr1': 'Owner_address'}, inplace=True)
if 'Nearby_owner' in dat.columns:
    k = 2
    dat = dat[['Target_fid', 'Status', 'Comments', 'Date', 'Submitted_by','Dogs_present', 'Hazardous_materials_present', 
    'Biohazards_present','Size_of_encampment', 'Maptaxlot_hyphen', 'Owner_name', 'Owner_address', 'Nearby_owner', 
    'Nearby_owner_address', 'Ownercity', 'Ownerprvst', 'Ownerzip', 'Geocity_name', 'Ugb_name','Longitude', 'Latitude']]
    taxlotcodes = pd.read_csv('mythical_taxlot_codes.csv')
    for idx in range(0, dat.shape[0]):
        taxlotcode = int(dat.loc[idx,'Maptaxlot_hyphen'][-2:])
        if np.isnan(dat.loc[idx, 'Owner_name']) & (taxlotcode in taxlotcodes.end_number.values):
            dat.loc[idx, 'Owner_name'] = 'Mythical lot number for ' + taxlotcodes.loc[taxlotcodes.end_number == taxlotcode, 'taxlot'].values[0]
else:
    k = 0

dat.rename(columns={'Target_fid': 'Target_FID'}, inplace=True)

intakepath = r'G:\projects\UtilityDistricts\eweb\DrinkingWater\RiparianEcosystemMarketplace\market_area\REM_area.gdb'
intake_areas = gpd.read_file(intakepath, driver='FileGDB', layer='AboveIntake')
points = gpd.read_file(path + '\\MyProject4.gdb', driver='FileGDB', layer='HomelessCampSite_SpatialJoin')
points = points.to_crs(epsg=2914)
print("Read data...")

urlstart ='https://services5.arcgis.com/9s1YtFmLS0YTl10F/arcgis/rest/services/ZHomeless_Camp_Trash_Collector/FeatureServer/0/'
urlend = '/attachments?f=html&token='

datestr = points.Date.values[0].split('T')[0]
res = convert_date(datestr)
Y = res[1]
m = res[2]
d = res[3]
outfolder = os.path.join(outpath, Y, m+'_'+d)
filename = 'IllegalCampNotice_'+m+'_'+d+'_'+Y[2:4]+'.xlsx'
file = os.path.join(outfolder, filename)
print("Got file name...")

for pID in points.index:
    point = points[points.index==pID]
    if all(intake_areas.contains(point)):
        dat.loc[pID, 'Above_Intake'] = 'Yes'
    else:
        dat.loc[pID, 'Above_Intake'] = 'No'
    FID = dat.loc[pID, 'Target_FID']
    url= urlstart+str(FID)+urlend
    resp=requests.get(url)
    links = []
    if resp.status_code==200:
        soup=BeautifulSoup(resp.text,'html.parser')
        for link in soup.findAll('a'):
            links.append(link.get('href'))
    else:
        print("Error")
    attached = [link for link in links if 'attachments' in link]
    if len(attached)==0:
        dat.loc[pID, 'Photos'] = 'NA'
    elif len(attached)==1:
        dat.loc[pID, 'Photos'] = 'https://services5.arcgis.com' + attached[0]
    else:
        dat.loc[pID, 'Photos'] = '; '.join(['https://services5.arcgis.com' + s for s in attached])

print("Edited data...")

dat.loc[:, 'Date'] = dat.Date.astype(str)
dat.to_excel(file, index=False)
print("Exported data...")

dat = pd.read_excel(file)
wb = openpyxl.load_workbook(file)
ws = wb.active
ws = removeFormatting(ws)
for pID in points.index:
    photovalue = dat.loc[pID, 'Photos']
    if photovalue == 'NA':
        print(f"No photos at Point {dat.loc[pID, 'Target_FID']}")
    else:
        ws.cell(row=2+pID, column=21+k).value = 'Yes'
        if ';' in photovalue: 
            urls = photovalue.split('; ')
            for i in range(0, len(urls)):
                ws.cell(row=2+pID, column=22+i+k).value = '=HYPERLINK("{}", "{}")'.format(urls[i],'Photo '+str(i+1))
                ws.cell(row=2+pID, column=22+i+k).style = "Hyperlink"
        else:
            ws.cell(row=2+pID, column=22+k).value = '=HYPERLINK("{}", "{}")'.format(photovalue,'Photo')
            ws.cell(row=2+pID, column=22+k).style = "Hyperlink"
wb.save(file)
print("Checked photos...")
excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(file)
excel.Worksheets(1).Activate()
excel.ActiveSheet.Columns.AutoFit()
wb.Close(True)
print("Autofitted columns...")


